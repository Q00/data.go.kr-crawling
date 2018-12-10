#-*- coding: utf-8 -*-
import config
import column
import openpyxl
import requests 
import urllib
from bs4 import BeautifulSoup as BS
from lxml import etree

#api key loading
key = config.go_data_api_key
#print(key)
#excel open
wb = openpyxl.Workbook()
ws = wb.active

#make cell head

#get Data
baseUrl = 'http://apis.data.go.kr/1470000/DURPrdlstInfoService/'
urlList = 'getDurPrdlstInfoList','getSeobangjeongPartitnAtentInfoList','getEfcyDplctInfoList','getOdsnAtentInfoList','getMdctnPdAtentInfoList','getCpctyAtentInfoList','getPwnmTabooInfoList','getSpcifyAgrdeTabooInfoList','getUsjntTabooInfoList'

for addUrl in urlList:
    print('검색할 url : ',addUrl) 
    #excel open
    wb = openpyxl.Workbook()
    ws = wb.active

    #excel parameter initialize
    main_row = 1 
    cell_num = 1

    #a~z loop
    for letter in range(ord('a'), ord('z')+1):
        params = {'itemName': chr(letter), 'Servicekey': key, 'numOfRows':100}
        if addUrl != 'getDurPrdlstInfoList':
            params.update({'typeName' : column.typeName[addUrl]})
        params_str = "&".join("%s=%s" % (k,v) for k,v in params.items())
        requesturl = baseUrl + addUrl
        #print(requesturl)
        try:
            #request data
            print('데이터 가져오는중 검색한 알파벳 :', chr(letter))
            getdata = requests.get(requesturl, params=params_str)
        except:
            print('request error')
        try:
            soup = BS(getdata.text, 'lxml-xml') 
            
            #check page
            totalCount = int(soup.find('totalCount').text)
            print('총 개수: ',totalCount)
            page = int(totalCount/100) + 1
            #cell number parameter
            for i in range(page):
                params_str2 = params_str
                params_str2+= '&pageNo='+str(i+1)
                #request again
                getdata = requests.get(requesturl, params=params_str2)
                soup = BS(getdata.text,'lxml-xml') 
                #print(soup.prettify)

                #validation check
                okflag = soup.find('resultCode')
                
                if okflag.text != '00':
                    print("okflag: ",okflag.text)
                    
                    raise ValueError('okcode is not 00')    
                else:
                    #검색이 잘 되었을때
                    for item in soup.find_all('item'):
                        #헤드 부분
                        if main_row ==1: 
                            for child in item.children:
                                print(child.name)
                                if child != '\n' and child.name != 'TYPE_NAME':
                                    evalue = column.typeList[addUrl][child.name]
                                    ws.cell(main_row,cell_num).value = evalue 
                                    cell_num+=1
                            main_row+=1

                        #cell number 초기화
                        cell_num =1

                        for child in item.children:
                            if child != '\n' and child.name != 'TYPE_NAME':
                                #print(child.contents)
                                #print(column.getDurPrdlstInfoList[child.name])
                                #엑셀 헤드 부분
                                evalue = str(child.text)

                                ws.cell(main_row,cell_num).value = evalue 
                                cell_num+=1
                        main_row+=1
        except ValueError as error:
            print(error) 
    print(column.typeName[addUrl]+'파일 완료')
    wb.save(column.typeName[addUrl]+'.xlsx')

